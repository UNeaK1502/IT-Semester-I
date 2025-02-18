import subprocess
import openpyxl
from openpyxl.styles import PatternFill
print("running compare.py")
def run_exe(exe_path, input_value):
    process = subprocess.Popen([exe_path], stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate(input=str(input_value).encode())
    return stdout.decode(errors='replace').strip()

def get_last_word(text):
    if isinstance(text, str):
        return text.split()[-1]
    return ''

def compare_exe_outputs(exe1_path, exe2_path, start, end, step):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Results"
    ws.append(["Input", "Output 1", "Output 2", "Difference"])
    for i in range(start, end + 1, step):
        output1 = run_exe(exe1_path, i)
        output2 = run_exe(exe2_path, i)
        last_word1 = get_last_word(output1)

        last_word2 = get_last_word(output2)
        difference = last_word1 != last_word2
        ws.append([i, output1, output2, difference])
        print("running number", i)
        if difference:
            for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=4):
                for c in cell:
                    c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save("comparison_results.xlsx")
    print("Comparison results saved to comparison_results.xlsx")

if __name__ == "__main__":
    exe1_path = r"C:\Users\UNeaK\Documents\Visual Studio 2022\repos\IT-Semester-I\x64\Release\Klausur-Semester-I.exe"
    exe2_path = r"C:\Users\UNeaK\Downloads\num2textso.exe"
    compare_exe_outputs(exe1_path, exe2_path, 900, 1000, 1)
