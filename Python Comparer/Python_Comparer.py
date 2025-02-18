import subprocess
import openpyxl
from openpyxl.styles import PatternFill
print("running compare.py")
def run_exe(exe_path, input_value):
    process = subprocess.Popen([exe_path], stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate(input=str(input_value).encode())
    return stdout.decode(errors='replace').strip()

def get_last_word(text):  
   if isinstance(text, str):  
       last_space_index = text.rfind(': ')     
       text = text[last_space_index + 1:] if last_space_index != -1 else text
       return text
   return ''

def compare_exe_outputs(exe1_path, exe2_path, start, end, step):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Results"
    ws.append(["Input", "Philip", "Steffen", "Difference"])
    for i in range(start, end + 1, step):
        output1 = run_exe(exe1_path, i)
        output2 = run_exe(exe2_path, i)
        last_word1 = get_last_word(output1)
        last_word2 = get_last_word(output2)
        output1 = last_word1
        output2 = last_word2
        difference = last_word1 != last_word2
        ws.append([i, output1, output2, difference])
        for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=4):
            if difference:
                for c in cell:
                    c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    wb.save("comparison_results.xlsx")
    print("Comparison results saved to comparison_results.xlsx")

if __name__ == "__main__":
    exe1_path = r"C:\Users\UNeaK\Documents\Visual Studio 2022\repos\IT-Semester-I\x64\Release\Klausur-Semester-I.exe"
    exe2_path = r"C:\Users\UNeaK\Downloads\num2textso.exe"      
    compare_exe_outputs(exe1_path, exe2_path, -1000,1000,1)
